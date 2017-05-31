# ToL generator program version 0
# ********************************************************************************************
# Revision          Date        Author          Reason/Summary of changes
# ----------------- ----------- --------------- ----------------------------------------------
# 0.0               8/28/2015   Zubin           Initial version, design process release
# 0.1               9/01/2015   Zubin           Added exception for "KeyError" on running out
#                                               of tabbed input.
# 0.2               9/08/2015   Zubin           Changed to command line arguments mode

# Notes on imports
# ****************************************************
# openpyxl package is the main package used in this  #
# program.                                           #
# ****************************************************

from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import datetime
import re
import sys

# Processing command line arguments
print('Number of args : {}', len(sys.argv))
if len(sys.argv) < 3:
    print("ERROR")
    exit()
print('arg 1 : ', str(sys.argv[0]))
print('arg 2 : ', str(sys.argv[1]))
print('arg 3 : ', str(sys.argv[2]))

# Set the parameters
row_entry = 2
# Input file parameter
# This will be a command line argment with absolute path, It is now changed to
# second argument in the command line.
# inputFile = "c:/Projects/Hosted MMSC/Working Notes/TOL_Input_b.xlsx"
inputFile = str(sys.argv[1])
wbookIN = load_workbook(inputFile)
parametersWS = 'Use case inputs'
paramWS = wbookIN[parametersWS]

# Output file & parameters
date_n_time = datetime.datetime.now()
fn_prefix = '%s%s%s_%s_%s_%s' % (date_n_time.year, date_n_time.month, date_n_time.day,
                    date_n_time.hour, date_n_time.minute, date_n_time.second)
# The output file has the same as the input file prefix
# The .xls/xlsx extention need to removes and time prefix need to be added
# outputFile = "c:/Projects/Hosted MMSC/Working Notes/TOL_Output_" + fn_prefix + ".xlsx" (CHANGED AS BELOW)
rpl_in = re.compile('(.xls|.xlsx)')
# outputFile = inputFile + fn_prefix + ".xlsx"
outputFile = rpl_in.sub('-', inputFile) + fn_prefix + ".xlsx"
# The template file is change to 2nd command line argument
# templatefile = "/Projects/Hosted MMSC/Working Notes/TOLTemplate_v1.xlsx" (CHANGED AS BELOW)
templatefile = str(sys.argv[2])
shutil.copy2(templatefile, outputFile)

wbookOUT = load_workbook(outputFile)
wsOUT = wbookOUT.active
# End of output file parameters


# Function definitions
# ********************************************************************
# Function name: is_tc_header, returns 1 if entry is a header entity
def is_tc_header(col):
    if col in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
        return 1
    else:
        return 0

# ********************************************************************
# Function name: write_test_header, writes header into an output TOL
def write_test_header(heading, col, rownumber):
    rowcolLocation = col + str(rownumber)
    wsOUT[rowcolLocation] = heading
    wbookOUT.save(filename=outputFile)

# ********************************************************************
# Function name: write_test_stepid, writes step ID into an output TOL
#                column number for step is set up 'G' column
#                rn is row number of the entry
def write_test_stepid(stepid, rownumber):
    rowcolLocation = 'G' + str(rownumber)
    wsOUT[rowcolLocation] = stepid
    wbookOUT.save(filename=outputFile)

# ******************************************************************************
# Function name: write_test_stepdesc, writes step description into an output TOL
#                column number for step description is set up 'H' column
#                rownumber is row number of the entry
def write_test_stepdesc(stepdesc, rownumber):
    rowcolLocation = 'H' + str(rownumber)
    wsOUT[rowcolLocation] = stepdesc
    wbookOUT.save(filename=outputFile)

# ******************************************************************************
# Function name: write_test_stepdesc, writes step description into an output TOL
#                column number for step description is set up 'H' column
#                rownumber is row number of the entry
def write_test_stepexpect(stepexpect, rownumber):
    rowcolLocation = 'I' + str(rownumber)
    wsOUT[rowcolLocation] = stepexpect
    wbookOUT.save(filename=outputFile)

# ******************************************************************************
# Function name: write_test_expected, writes expected result into an output TOL
#                column number for step description is set up 'I' column
#                rownumber is row number of the entry
def write_test_expected(expresult, rownumber):
    rowcolLocation = 'I' + str(rownumber)
    wsOUT[rowcolLocation] = expresult
    wbookOUT.save(filename=outputFile)

# ******************************************************************************
# Function name: create_step_from_dict, return a steps text built from
#                dictionary data available from input values
#                tcstepstr is the string input from the test case seed
#                td_dictionary is the test data dictionary
def create_step_from_dict(tcstepstr, td_dictionary):
    for repl_match, repl_value in td_dictionary.items():
        tcstepstr = tcstepstr.replace(repl_match, repl_value)
    return tcstepstr

# End of function definitions.

# Iterating the input parameter file

argNames = []
argValues = []
# ********************************************************************************
# Notes: Begining of the main "for loop"
# This is the main for loop for reading the input work sheet. Each iteration reads
# the row till the last non empty column. The first row is replacement dictionary
# there  for not  processed, however  the  values are  stored as  the replacement
# dictionary in the array argNames.
# ********************************************************************************

for row in paramWS.rows:
    dict_list = {}
    for cell in row:
        currentRow = cell.row
        currentCol = cell.column
        if currentRow != 1:
            argValues.append(cell.value)
            currArgumentLocation = currentCol + "1"
            currArgumentName = paramWS[currArgumentLocation].value
            currArgument = cell.value
            # Check if it is a new test case
            if is_tc_header(currentCol):
                # Adding the header entry for a test case
                if currentCol == 'A':
                    write_test_header(currArgument, currentCol, row_entry)
                if currentCol == 'B':
                    write_test_header(currArgument, currentCol, row_entry)
                if currentCol == 'C':
                    write_test_header(currArgument, currentCol, row_entry)
                if currentCol == 'D':
                    write_test_header(currArgument, currentCol, row_entry)
                if currentCol == 'E':
                    write_test_header(currArgument, currentCol, row_entry)
                if currentCol == 'F':
                    write_test_header(currArgument, currentCol, row_entry)

            # Building a dictionary pair, excluding headers in dictionary
            if currentCol not in ('A', 'B', 'C', 'D', 'E', 'F'):
                dict_list[currArgumentName] = str(currArgument)

            # Get the sheet where the test generic seed is located
            # Note this field has to be the last in the input parameter spread sheet
            #print(currArgumentName)
            if re.match('<testtab>', currArgumentName):
                testCaseTAB = currArgument
                # The following try-except was added in order to trap the "KeyError"
                # exceptions that arises on the last input line is surpassed.
                try:
                    tcWS = wbookIN[testCaseTAB]
                except KeyError:
                    #print("KeyError: possible end of input")
                    exit()
                for tcrow in tcWS.rows:
                    for tccell in tcrow:
                        if tccell.row != 1:
                            if tccell.column == 'A':
                                write_test_stepid(tccell.value, row_entry)
                            if tccell.column == 'B':
                                tc_step_string = tccell.value
                                tc_step_built = create_step_from_dict(tc_step_string, dict_list)
                                write_test_stepdesc(tc_step_built, row_entry)
                            if tccell.column == 'C':
                                tc_step_string = tccell.value
                                tc_step_built = create_step_from_dict(tc_step_string, dict_list)
                                write_test_stepexpect(tc_step_built, row_entry)
                    # Incrementing to the next row for the next set of arguments.
                    row_entry = row_entry + 1
        else:
            argNames.append(cell.value)

    # End of the inner column travesing, left to right loop.
# End of the main for loop, traversing rows from top to bottom.